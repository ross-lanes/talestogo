from dotenv import load_dotenv
load_dotenv(override=True)  # Load environment variables from .env file

from fastapi import FastAPI, Depends, HTTPException, status, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import datetime
from datetime import timedelta
import os
import subprocess
import io
import base64
from pathlib import Path

# Use explicit imports from the 'app' package
from app import crud, models, schemas
from app.database import SessionLocal, engine
from app.routers import analytics, admin
from app.auth import (
    get_current_user,
    get_current_admin_user,
    authenticate_user,
    create_access_token,
    get_password_hash,
    encrypt_api_key,
    decrypt_api_key,
    verify_google_token,
    verify_microsoft_token,
    get_or_create_oauth_user,
    ACCESS_TOKEN_EXPIRE_MINUTES
)

# This line ensures tables are created if they don't exist when the app starts.
# It's convenient for development. For production, you'd use a migration tool.
models.Base.metadata.create_all(bind=engine)

# Create the FastAPI app instance
app = FastAPI(
    title="Tales",
    description="An AI tool for tracking and analyzing LLM brand depictions.",
    version="0.1.0"
)

# Configure CORS to allow frontend to access backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",  # Local development
        "https://tales-frontend.onrender.com",  # Production frontend (legacy)
        "https://tales.robotrachel.com"  # Production frontend (custom domain)
    ],
    allow_credentials=True,
    allow_methods=["*"],  # Allow all HTTP methods
    allow_headers=["*"],  # Allow all headers
)

# Include routers
app.include_router(analytics.router)
app.include_router(admin.router)

# --- Dependencies ---
def get_db():
    """
    Dependency function that provides a database session per request.
    Ensures the session is always closed, even if errors occur.
    """
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_active_brand_id(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> Optional[int]:
    """
    Helper function to get the active brand_id for the current user.
    Returns None if no active brand exists (allows multi-brand view).
    """
    active_brand = crud.get_active_brand(db, user_id=current_user.id)
    return active_brand.id if active_brand else None

# --- Root Endpoint ---
@app.get("/", tags=["Root"])
async def read_root():
    """Root endpoint to check if the API is running."""
    return {"message": "Welcome to the TALES API!"}


# --- Authentication Endpoints ---
@app.post("/auth/register", response_model=schemas.User, status_code=201, tags=["Authentication"])
def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    """
    Register a new user (invite-only).
    User account will be inactive until admin approves.
    """
    # Check if user already exists
    db_user = crud.get_user_by_email(db, email=user.email)
    if db_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )

    # Hash password and create user
    hashed_password = get_password_hash(user.password)
    new_user = crud.create_user(db=db, user=user, hashed_password=hashed_password, is_invited=False)
    return new_user


@app.post("/auth/login", response_model=schemas.Token, tags=["Authentication"])
def login(user_credentials: schemas.UserLogin, db: Session = Depends(get_db)):
    """
    Login with email and password.
    Returns JWT access token.
    """
    user = authenticate_user(db, email=user_credentials.email, password=user_credentials.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account is not active. Please contact admin for approval."
        )

    # Create access token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": str(user.id)},
        expires_delta=access_token_expires
    )

    return {"access_token": access_token, "token_type": "bearer"}


@app.post("/auth/google", response_model=schemas.Token, tags=["Authentication"])
def google_login(google_token: schemas.GoogleLogin, db: Session = Depends(get_db)):
    """
    Login with Google OAuth.
    Accepts Google ID token and returns JWT access token.
    Creates new user if doesn't exist (auto-activated).
    """
    # Verify Google token and get user info
    google_info = verify_google_token(google_token.token)

    # Ensure email is verified
    if not google_info.get('email_verified'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email not verified with Google"
        )

    # Get or create user
    user = get_or_create_oauth_user(db, google_info)

    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account is not active. Please contact admin for approval."
        )

    # Create access token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": str(user.id)},
        expires_delta=access_token_expires
    )

    return {"access_token": access_token, "token_type": "bearer"}


@app.post("/auth/microsoft", response_model=schemas.Token, tags=["Authentication"])
def microsoft_login(microsoft_token: schemas.MicrosoftLogin, db: Session = Depends(get_db)):
    """
    Login with Microsoft OAuth.
    Accepts Microsoft ID token and returns JWT access token.
    Creates new user if doesn't exist (auto-activated for admin).
    """
    # Verify Microsoft token and get user info
    microsoft_info = verify_microsoft_token(microsoft_token.token)

    # Ensure email is verified
    if not microsoft_info.get('email_verified'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email not verified with Microsoft"
        )

    # Get or create user
    user = get_or_create_oauth_user(db, microsoft_info, provider='microsoft')

    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account is not active. Please contact admin for approval."
        )

    # Create access token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": str(user.id)},
        expires_delta=access_token_expires
    )

    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/auth/me", response_model=schemas.User, tags=["Authentication"])
def get_current_user_info(current_user: models.User = Depends(get_current_user)):
    """Get current logged-in user information."""
    return current_user


@app.put("/auth/me", response_model=schemas.User, tags=["Authentication"])
def update_current_user_profile(
    user_update: schemas.UserUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update current user's profile and API keys."""
    updated_user = crud.update_user(
        db,
        user_id=current_user.id,
        user_update=user_update,
        encrypt_keys_func=encrypt_api_key
    )
    if not updated_user:
        raise HTTPException(status_code=404, detail="User not found")
    return updated_user


# --- Admin User Management Endpoints ---
@app.get("/admin/users", response_model=List[schemas.User], tags=["Admin"])
def list_all_users(
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """List all users (admin only)."""
    return crud.get_users(db, skip=skip, limit=limit)


@app.post("/admin/users/invite", response_model=schemas.User, status_code=201, tags=["Admin"])
def invite_user(
    user_invite: schemas.UserInvite,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """
    Create an invited user (admin only).
    Invited users are pre-approved (is_active=True).
    """
    # Check if user already exists
    db_user = crud.get_user_by_email(db, email=user_invite.email)
    if db_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )

    # Create temporary password (user should change on first login)
    temp_password = "TempPassword123!"
    hashed_password = get_password_hash(temp_password)

    # Create user object
    user_create = schemas.UserCreate(
        email=user_invite.email,
        password=temp_password,
        full_name=user_invite.full_name,
        organization=user_invite.organization
    )

    new_user = crud.create_user(db=db, user=user_create, hashed_password=hashed_password, is_invited=True)

    # Activate invited user immediately
    admin_update = schemas.UserAdminUpdate(is_active=True)
    crud.admin_update_user(db, user_id=new_user.id, user_update=admin_update)

    return new_user


@app.put("/admin/users/{user_id}", response_model=schemas.User, tags=["Admin"])
def admin_update_user_status(
    user_id: int,
    user_update: schemas.UserAdminUpdate,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Update user status (activate/deactivate, make admin) - admin only."""
    updated_user = crud.admin_update_user(db, user_id=user_id, user_update=user_update)
    if not updated_user:
        raise HTTPException(status_code=404, detail="User not found")
    return updated_user


@app.delete("/admin/users/{user_id}", response_model=schemas.User, tags=["Admin"])
def admin_delete_user(
    user_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Delete a user (admin only). Cannot delete yourself."""
    # Prevent admin from deleting themselves
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You cannot delete your own account"
        )

    # Get the user
    user_to_delete = db.query(models.User).filter(models.User.id == user_id).first()
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")

    # Delete the user (will cascade delete all related data)
    db.delete(user_to_delete)
    db.commit()

    return user_to_delete


# --- Invitation Endpoints ---
@app.post("/admin/users/create-invite", response_model=schemas.InvitationResponse, status_code=201, tags=["Admin", "Invitations"])
def create_invitation(
    invitation: schemas.InvitationCreate,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """
    Add user email to approved list (admin only).
    User can then login directly with Google OAuth.
    No invitation token needed - just tell them to visit the site and sign in with Google.
    """
    # Check if user already exists
    existing_user = crud.get_user_by_email(db, email=invitation.email)
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User with this email already exists"
        )

    # Create pre-approved user (active and ready for OAuth login)
    new_user = models.User(
        email=invitation.email,
        full_name=invitation.full_name,
        is_invited=True,
        is_active=True,  # Pre-approved, will be activated on first Google login
        invitation_token=None,
        invitation_expires_at=None
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    # Return simple login URL
    frontend_url = os.getenv("FRONTEND_URL", "http://localhost:5173")

    return schemas.InvitationResponse(
        email=invitation.email,
        full_name=invitation.full_name,
        invitation_token="",  # No token needed
        expires_at=None,
        invitation_url=frontend_url  # Just send them to the main site
    )


@app.get("/invite/validate", response_model=schemas.InvitationValidate, tags=["Invitations"])
def validate_invitation(token: str, db: Session = Depends(get_db)):
    """
    Validate an invitation token.
    Returns user info if token is valid, otherwise raises error.
    """
    from app.auth import verify_invitation_token

    # Verify token
    payload = verify_invitation_token(token)

    # Check if user exists and invitation is still valid
    user = crud.get_user_by_email(db, email=payload["email"])
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invitation not found"
        )

    if user.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This invitation has already been used"
        )

    if user.invitation_token != token:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid invitation token"
        )

    if user.invitation_expires_at and user.invitation_expires_at < datetime.datetime.utcnow():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This invitation has expired"
        )

    return schemas.InvitationValidate(
        email=payload["email"],
        full_name=payload["full_name"],
        expires_at=user.invitation_expires_at
    )


@app.post("/invite/accept", response_model=schemas.Token, tags=["Invitations"])
def accept_invitation(
    invitation_accept: schemas.InvitationAccept,
    db: Session = Depends(get_db)
):
    """
    Accept an invitation by setting password and activating account.
    Returns access token for immediate login.
    """
    from app.auth import verify_invitation_token, get_password_hash, create_access_token

    # Verify token
    payload = verify_invitation_token(invitation_accept.token)

    # Get user
    user = crud.get_user_by_email(db, email=payload["email"])
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invitation not found"
        )

    if user.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This invitation has already been used"
        )

    if user.invitation_token != invitation_accept.token:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid invitation token"
        )

    if user.invitation_expires_at and user.invitation_expires_at < datetime.datetime.utcnow():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This invitation has expired"
        )

    # Set password and activate user
    user.hashed_password = get_password_hash(invitation_accept.password)
    user.is_active = True
    user.invitation_token = None  # Clear token after use
    db.commit()
    db.refresh(user)

    # Create access token for immediate login
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": str(user.id)},
        expires_delta=access_token_expires
    )

    return {"access_token": access_token, "token_type": "bearer"}


# --- Query Endpoints ---
@app.post("/queries/", response_model=schemas.Query, status_code=201, tags=["Queries"])
def create_query(
    query: schemas.QueryCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Create a new query for the current user's active brand."""
    db_query = crud.get_query_by_query_id(db, query_id=query.query_id, user_id=current_user.id, brand_id=brand_id)
    if db_query:
        raise HTTPException(status_code=400, detail=f"Query ID {query.query_id} already exists for this brand")
    return crud.create_query(db=db, query=query, user_id=current_user.id, brand_id=brand_id)

@app.get("/queries/", response_model=List[schemas.Query], tags=["Queries"])
def read_queries(
    active_only: bool = False,
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of queries for the current user's active brand."""
    if active_only:
        queries = crud.get_active_queries(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)
    else:
        queries = crud.get_queries(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)
    return queries

@app.get("/queries/{query_id}", response_model=schemas.Query, tags=["Queries"])
def read_query(
    query_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a single query by its user-facing ID (e.g., 'Q001') for the current user's active brand."""
    db_query = crud.get_query_by_query_id(db, query_id=query_id, user_id=current_user.id, brand_id=brand_id)
    if db_query is None:
        raise HTTPException(status_code=404, detail="Query not found")
    return db_query

@app.put("/queries/{query_id}", response_model=schemas.Query, tags=["Queries"])
def update_query(
    query_id: str,
    query_update: schemas.QueryUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a query for the current user's active brand."""
    db_query = crud.update_query(db, query_id=query_id, query_update=query_update, user_id=current_user.id, brand_id=brand_id)
    if db_query is None:
        raise HTTPException(status_code=404, detail="Query not found")
    return db_query

@app.delete("/queries/{query_id}", response_model=schemas.Query, tags=["Queries"])
def delete_query(
    query_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a query for the current user's active brand."""
    deleted_query = crud.delete_query(db, query_id=query_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_query is None:
        raise HTTPException(status_code=404, detail="Query not found")
    return deleted_query

@app.post("/queries/upload", tags=["Queries"])
async def upload_queries(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Upload queries from an Excel file (admin only). Expected columns: query_id, query_text, category, target_outcome, brand_in_query, active"""
    from openpyxl import load_workbook

    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand selected")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    try:
        # Read the Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Validate required columns
        required_columns = ['query_id', 'query_text']
        for col in required_columns:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        created_count = 0
        updated_count = 0
        errors = []

        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dict from row
                row_data = dict(zip(headers, row))

                if not row_data.get('query_id'):
                    continue  # Skip empty rows

                # Check if query already exists
                existing_query = db.query(models.Query).filter(
                    models.Query.query_id == row_data['query_id'],
                    models.Query.user_id == current_user.id,
                    models.Query.brand_id == brand_id
                ).first()

                if existing_query:
                    # Update existing query
                    existing_query.query_text = row_data.get('query_text', existing_query.query_text)
                    existing_query.category = row_data.get('category', existing_query.category)
                    existing_query.target_outcome = row_data.get('target_outcome', existing_query.target_outcome)
                    existing_query.brand_in_query = row_data.get('brand_in_query', existing_query.brand_in_query)
                    if 'active' in row_data and row_data['active'] is not None:
                        existing_query.active = bool(row_data['active'])
                    updated_count += 1
                else:
                    # Create new query
                    new_query = models.Query(
                        user_id=current_user.id,
                        brand_id=brand_id,
                        query_id=row_data['query_id'],
                        query_text=row_data['query_text'],
                        category=row_data.get('category'),
                        target_outcome=row_data.get('target_outcome'),
                        brand_in_query=row_data.get('brand_in_query', False),
                        active=row_data.get('active', True) if row_data.get('active') is not None else True
                    )
                    db.add(new_query)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        db.commit()

        return {
            "message": "Queries uploaded successfully",
            "created": created_count,
            "updated": updated_count,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")


# --- Response Endpoints ---
@app.post("/responses/", response_model=schemas.Response, status_code=201, tags=["Responses"])
def create_response(
    response: schemas.ResponseCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Submit a raw response from an LLM platform for the current user's active brand."""
    return crud.create_response(db=db, response=response, user_id=current_user.id, brand_id=brand_id)

@app.get("/responses/", response_model=List[schemas.Response], tags=["Responses"])
def read_responses(
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of responses for the current user's active brand."""
    return crud.get_responses(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)

@app.get("/responses/unanalyzed/", response_model=List[schemas.Response], tags=["Responses"])
def read_unanalyzed_responses(
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Retrieve responses that are pending analysis for the current user."""
    return crud.get_unanalyzed_responses(db, user_id=current_user.id, limit=limit)

@app.put("/responses/{response_id}/analyze", response_model=schemas.Response, tags=["Responses"])
def update_response_analysis(
    response_id: int,
    analysis_data: schemas.ResponseAnalysisInput,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a response with analysis data for the current user's active brand."""
    db_response = crud.update_response_analysis(
        db,
        response_id=response_id,
        analysis_data=analysis_data.model_dump(exclude_unset=True),
        user_id=current_user.id,
        brand_id=brand_id
    )
    if db_response is None:
        raise HTTPException(status_code=404, detail="Response not found")
    return db_response

@app.delete("/responses/{response_id}", response_model=schemas.Response, tags=["Responses"])
def delete_response(
    response_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a response for the current user's active brand."""
    deleted_response = crud.delete_response(db, response_id=response_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_response is None:
        raise HTTPException(status_code=404, detail="Response not found")
    return deleted_response

@app.get("/responses/export/excel", tags=["Responses"])
def export_responses_to_excel(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Export all responses for the active brand to an Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    # Get all responses for the active brand
    responses = crud.get_responses(db, user_id=current_user.id, brand_id=brand_id, skip=0, limit=10000)

    if not responses:
        raise HTTPException(status_code=404, detail="No responses found for export")

    # Get brand name for filename
    brand = crud.get_active_brand(db, user_id=current_user.id) if brand_id else None
    brand_name = brand.brand_name if brand else "Unknown"

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Responses"

    # Define headers
    headers = [
        "Response ID",
        "Query ID",
        "Platform",
        "Response Text",
        "Collected At",
        "Analyzed At",
        "Brand Mentioned",
        "Brand Position",
        "Sentiment",
        "Descriptors",
        "Competitors",
        "Sources",
        "Notes"
    ]

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for row_num, response in enumerate(responses, 2):
        ws.cell(row=row_num, column=1).value = response.id
        ws.cell(row=row_num, column=2).value = response.query_id
        ws.cell(row=row_num, column=3).value = response.platform
        ws.cell(row=row_num, column=4).value = response.response_text
        ws.cell(row=row_num, column=5).value = response.timestamp.strftime('%Y-%m-%d %H:%M:%S') if response.timestamp else ''
        ws.cell(row=row_num, column=6).value = response.analyzed_at.strftime('%Y-%m-%d %H:%M:%S') if response.analyzed_at else ''
        ws.cell(row=row_num, column=7).value = response.brand_mentioned or ''
        ws.cell(row=row_num, column=8).value = response.brand_position or ''
        ws.cell(row=row_num, column=9).value = response.sentiment or ''
        ws.cell(row=row_num, column=10).value = response.descriptors or ''
        ws.cell(row=row_num, column=11).value = response.competitors or ''
        ws.cell(row=row_num, column=12).value = response.sources or ''
        ws.cell(row=row_num, column=13).value = response.notes or ''

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create safe filename
    safe_brand_name = "".join(c for c in brand_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"{safe_brand_name}_AI_Responses.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# --- Competitor Endpoints ---
@app.post("/competitors/", response_model=schemas.Competitor, status_code=201, tags=["Competitors"])
def create_competitor(
    competitor: schemas.CompetitorCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Create a new competitor for the current user's active brand."""
    return crud.create_competitor(db=db, competitor=competitor, user_id=current_user.id, brand_id=brand_id)

@app.get("/competitors/", response_model=List[schemas.Competitor], tags=["Competitors"])
def read_competitors(
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of competitors for the current user's active brand."""
    return crud.get_competitors(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)

@app.put("/competitors/{competitor_id}", response_model=schemas.Competitor, tags=["Competitors"])
def update_competitor(
    competitor_id: int,
    competitor_update: schemas.CompetitorUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a competitor for the current user's active brand."""
    db_competitor = crud.update_competitor(
        db,
        competitor_id=competitor_id,
        competitor_update=competitor_update,
        user_id=current_user.id,
        brand_id=brand_id
    )
    if db_competitor is None:
        raise HTTPException(status_code=404, detail="Competitor not found")
    return db_competitor

@app.delete("/competitors/{competitor_id}", response_model=schemas.Competitor, tags=["Competitors"])
def delete_competitor(
    competitor_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a competitor for the current user's active brand."""
    deleted_competitor = crud.delete_competitor(db, competitor_id=competitor_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_competitor is None:
        raise HTTPException(status_code=404, detail="Competitor not found")
    return deleted_competitor

@app.post("/competitors/upload", tags=["Competitors"])
async def upload_competitors(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Upload competitors from an Excel file (admin only). Expected columns: competitor_name, active"""
    from openpyxl import load_workbook

    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand selected")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    try:
        # Read the Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Validate required columns
        required_columns = ['competitor_name']
        for col in required_columns:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        created_count = 0
        updated_count = 0
        errors = []

        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dict from row
                row_data = dict(zip(headers, row))

                if not row_data.get('competitor_name'):
                    continue  # Skip empty rows

                # Check if competitor already exists
                existing_competitor = db.query(models.Competitor).filter(
                    models.Competitor.competitor_name == row_data['competitor_name'],
                    models.Competitor.user_id == current_user.id,
                    models.Competitor.brand_id == brand_id
                ).first()

                if existing_competitor:
                    # Update existing competitor
                    if 'active' in row_data and row_data['active'] is not None:
                        existing_competitor.active = bool(row_data['active'])
                    updated_count += 1
                else:
                    # Create new competitor
                    new_competitor = models.Competitor(
                        user_id=current_user.id,
                        brand_id=brand_id,
                        competitor_name=row_data['competitor_name'],
                        active=row_data.get('active', True) if row_data.get('active') is not None else True
                    )
                    db.add(new_competitor)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        db.commit()

        return {
            "message": "Competitors uploaded successfully",
            "created": created_count,
            "updated": updated_count,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")


# --- Descriptor Endpoints ---
@app.post("/descriptors/", response_model=schemas.TargetDescriptor, status_code=201, tags=["Descriptors"])
def create_descriptor(
    descriptor: schemas.TargetDescriptorCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Create a new descriptor for the current user's active brand."""
    return crud.create_descriptor(db=db, descriptor=descriptor, user_id=current_user.id, brand_id=brand_id)

@app.get("/descriptors/", response_model=List[schemas.TargetDescriptor], tags=["Descriptors"])
def read_descriptors(
    target_only: bool = False,
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of descriptors for the current user's active brand."""
    if target_only:
        return crud.get_target_descriptors(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)
    return crud.get_descriptors(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)

@app.get("/descriptors/{descriptor_id}", response_model=schemas.TargetDescriptor, tags=["Descriptors"])
def read_descriptor(
    descriptor_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a single descriptor by its primary key for the current user's active brand."""
    db_descriptor = crud.get_descriptor(db, descriptor_id=descriptor_id, user_id=current_user.id, brand_id=brand_id)
    if db_descriptor is None:
        raise HTTPException(status_code=404, detail="Descriptor not found")
    return db_descriptor

@app.put("/descriptors/{descriptor_id}", response_model=schemas.TargetDescriptor, tags=["Descriptors"])
def update_descriptor(
    descriptor_id: int,
    descriptor_update: schemas.TargetDescriptorUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a descriptor for the current user's active brand."""
    db_descriptor = crud.update_descriptor(
        db,
        descriptor_id=descriptor_id,
        descriptor_update=descriptor_update,
        user_id=current_user.id,
        brand_id=brand_id
    )
    if db_descriptor is None:
        raise HTTPException(status_code=404, detail="Descriptor not found")
    return db_descriptor

@app.delete("/descriptors/{descriptor_id}", response_model=schemas.TargetDescriptor, tags=["Descriptors"])
def delete_descriptor(
    descriptor_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a descriptor for the current user's active brand."""
    deleted_descriptor = crud.delete_descriptor(db, descriptor_id=descriptor_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_descriptor is None:
        raise HTTPException(status_code=404, detail="Descriptor not found")
    return deleted_descriptor

@app.post("/descriptors/upload", tags=["Descriptors"])
async def upload_descriptors(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Upload descriptors from an Excel file (admin only). Expected columns: descriptor, is_target, current_ownership, priority, notes"""
    from openpyxl import load_workbook

    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand selected")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    try:
        # Read the Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Validate required columns
        required_columns = ['descriptor']
        for col in required_columns:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        created_count = 0
        updated_count = 0
        errors = []

        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dict from row
                row_data = dict(zip(headers, row))

                if not row_data.get('descriptor'):
                    continue  # Skip empty rows

                # Check if descriptor already exists
                existing_descriptor = db.query(models.TargetDescriptor).filter(
                    models.TargetDescriptor.descriptor == row_data['descriptor'],
                    models.TargetDescriptor.user_id == current_user.id,
                    models.TargetDescriptor.brand_id == brand_id
                ).first()

                if existing_descriptor:
                    # Update existing descriptor
                    if 'is_target' in row_data and row_data['is_target'] is not None:
                        existing_descriptor.is_target = bool(row_data['is_target'])
                    if 'current_ownership' in row_data:
                        existing_descriptor.current_ownership = row_data.get('current_ownership')
                    if 'priority' in row_data:
                        existing_descriptor.priority = row_data.get('priority')
                    if 'notes' in row_data:
                        existing_descriptor.notes = row_data.get('notes')
                    updated_count += 1
                else:
                    # Create new descriptor
                    new_descriptor = models.TargetDescriptor(
                        user_id=current_user.id,
                        brand_id=brand_id,
                        descriptor=row_data['descriptor'],
                        is_target=row_data.get('is_target', True) if row_data.get('is_target') is not None else True,
                        current_ownership=row_data.get('current_ownership'),
                        priority=row_data.get('priority', 'Medium'),
                        notes=row_data.get('notes')
                    )
                    db.add(new_descriptor)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        db.commit()

        return {
            "message": "Descriptors uploaded successfully",
            "created": created_count,
            "updated": updated_count,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")


# --- Report Endpoints ---
@app.post("/reports/", response_model=schemas.Report, status_code=201, tags=["Reports"])
def create_report(
    report: schemas.ReportCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new report for the current user."""
    return crud.create_report(db=db, report=report, user_id=current_user.id)

@app.get("/reports/", response_model=List[schemas.Report], tags=["Reports"])
def read_reports(
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of reports for the current user's active brand."""
    return crud.get_reports(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)

@app.get("/reports/{report_id}", response_model=schemas.Report, tags=["Reports"])
def read_report(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Retrieve a single report by its ID for the current user."""
    db_report = crud.get_report(db, report_id=report_id, user_id=current_user.id)
    if db_report is None:
        raise HTTPException(status_code=404, detail="Report not found")
    return db_report

@app.put("/reports/{report_id}", response_model=schemas.Report, tags=["Reports"])
def update_report(
    report_id: int,
    report_update: schemas.ReportUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a report (mainly for adding Google Doc URL) for the current user."""
    db_report = crud.update_report(
        db,
        report_id=report_id,
        report_update=report_update,
        user_id=current_user.id
    )
    if db_report is None:
        raise HTTPException(status_code=404, detail="Report not found")
    return db_report


@app.delete("/reports/{report_id}", response_model=schemas.Report, tags=["Reports"])
def delete_report(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a report for the current user."""
    deleted_report = crud.delete_report(db, report_id=report_id, user_id=current_user.id)
    if deleted_report is None:
        raise HTTPException(status_code=404, detail="Report not found")
    return deleted_report


@app.post("/reports/upload-charts", tags=["Reports"])
async def upload_chart_images(
    dashboard: Optional[str] = Form(None),
    sentiment: Optional[str] = Form(None),
    positioning: Optional[str] = Form(None),
    share_of_voice: Optional[str] = Form(None),
    timestamp: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """
    Upload chart images (as base64 strings) and save them to disk with a predictable naming pattern.
    These will be used by report generation if available.
    """
    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand found")

    # Create report_charts directory if it doesn't exist
    charts_dir = Path("frontend/public/report_charts")
    charts_dir.mkdir(parents=True, exist_ok=True)

    chart_paths = {}

    # Process each chart image with predictable naming
    charts_data = {
        'dashboard': dashboard,
        'sentiment': sentiment,
        'positioning': positioning,
        'share_of_voice': share_of_voice
    }

    for chart_name, base64_data in charts_data.items():
        if base64_data:
            try:
                # Remove the data:image/png;base64, prefix if present
                if ',' in base64_data:
                    base64_data = base64_data.split(',')[1]

                # Decode base64 to binary
                image_data = base64.b64decode(base64_data)

                # Create filename with predictable pattern that report generation can find
                # Format: {user_id}_{brand_id}_latest_{chart_name}.png
                filename = f"{current_user.id}_{brand_id}_latest_{chart_name}.png"
                filepath = charts_dir / filename

                # Write file
                with open(filepath, 'wb') as f:
                    f.write(image_data)

                # Store relative path for markdown
                chart_paths[chart_name] = f"report_charts/{filename}"

            except Exception as e:
                print(f"Error saving {chart_name} chart: {e}")
                continue

    return {
        "success": True,
        "chart_paths": chart_paths,
        "message": f"Uploaded {len(chart_paths)} chart(s) for report generation"
    }


@app.get("/reports/{report_id}/export/word", tags=["Reports"])
def export_report_to_word(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Export a report to Word document format with embedded charts."""
    from app.services.report_export import export_to_word_with_charts

    # Get the report
    db_report = crud.get_report(db, report_id=report_id, user_id=current_user.id)
    if db_report is None:
        raise HTTPException(status_code=404, detail="Report not found")

    # Generate Word document with charts
    word_file = export_to_word_with_charts(
        db_report.report_content,
        db_report.title,
        db,
        user_id=current_user.id,
        brand_id=brand_id
    )

    # Create safe filename
    safe_filename = "".join(c for c in db_report.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"{safe_filename}.docx"

    return StreamingResponse(
        word_file,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/reports/{report_id}/export/pdf", tags=["Reports"])
def export_report_to_pdf(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export a report to PDF format."""
    from app.services.report_export import export_to_pdf

    # Get the report
    db_report = crud.get_report(db, report_id=report_id, user_id=current_user.id)
    if db_report is None:
        raise HTTPException(status_code=404, detail="Report not found")

    # Generate PDF
    pdf_file = export_to_pdf(db_report.report_content, db_report.title)

    # Create safe filename
    safe_filename = "".join(c for c in db_report.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"{safe_filename}.pdf"

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/export/how-tales-works/word", tags=["Export"])
def export_how_tales_works_word(
    current_user: models.User = Depends(get_current_user)
):
    """Export How Tales Works methodology page as Word document."""
    from app.services.report_export import export_to_word

    # Methodology content
    methodology_content = """# How Tales Works

## Data Collection Methods

The analysis of your brand's AI reputation is conducted using the Tales platform, which employs a systematic multi-platform AI querying methodology. Data is collected by submitting strategically designed queries to four major large language model platforms: ChatGPT (OpenAI GPT-4o), Claude (Anthropic Claude-3-Haiku), Gemini (Google Gemini 2.5 Flash), and Perplexity (Sonar model with web search capabilities). These queries are automatically generated using AI to cover relevant topic areas including leadership and reputation, technology and innovation, and industry positioning.

Critically, most queries are designed as "visibility tests" that deliberately exclude your brand name, allowing the study to measure organic mentions—instances where AI platforms independently reference your organization when answering relevant questions. A list of descriptive words ideally included in AI responses to describe your brand are also automatically generated by AI, as well as a list of your competitors. The queries, descriptors and competitors are all reviewed by a human and edited to ensure they fit your brand's needs. All responses are collected via API with timestamps and platform metadata, enabling temporal and cross-platform comparative analysis.

**Note:** While the Tales platform allows users to assign priority levels (High, Medium, Low) to queries and descriptors for organizational purposes, these priority designations do not impact the quantitative analysis or metric calculations in any way—all queries and descriptors are weighted equally in the analysis.

## Analytical Framework

The collected responses undergo a two-stage analysis process combining structured data extraction with AI-powered insight generation. In the first stage, Perplexity's Sonar model analyzes each response to extract structured data including mention type (direct, indirect, or absent), brand positioning (categorized as leader, top 3, featured, listed, or not mentioned), sentiment classification (very positive, positive, neutral, negative, or mixed), associated descriptors and adjectives, competitor mentions, and cited sources.

This extraction process is context-aware, incorporating your brand's industry context, strategic messaging, target descriptors, and known competitors to ensure relevant and accurate classification. In the second stage, Perplexity's Sonar Pro model synthesizes these structured findings with real-time industry news and comprehensive brand context to generate strategic insights and actionable recommendations, explicitly connecting each finding to specific performance gaps and opportunities.

## Key Performance Metrics

The study calculates multiple quantitative metrics to assess your brand's AI reputation:

- **Mention Rate:** Measures the percentage of responses where the brand is referenced when not explicitly included in the query, indicating organic visibility in AI responses.
- **Sentiment Distribution:** Tracks the breakdown of positive, neutral, and negative associations across all mentions.
- **Brand Positioning:** Analyzes where your brand appears in AI-generated lists and discussions, calculating an average positioning score and leadership visibility percentage.
- **Descriptor Match Rate:** Compares target descriptors that your brand aims to be associated with against descriptors actually used by AI platforms, identifying alignment gaps.
- **Share of Voice:** Quantifies your brand's mentions relative to total organization mentions across all responses, including competitors, with weighting based on positioning strength.

All metrics are segmented by platform to identify which AI systems perform better or worse for your brand.

## Mathematical Formulas for Metric Calculations

### Brand Mentions (Mention Rate)

The mention rate quantifies how frequently your brand is referenced by AI platforms when the brand name is not explicitly included in the query, measuring organic visibility.

**Formula:**
```
Mention Rate (%) = (Number of Mentions / Total Qualifying Responses) × 100
```

| Component | Definition |
|-----------|------------|
| Numerator | Count of responses where brand_mentioned field equals 'Yes' OR 'Indirect' |
| Denominator | Total count of all responses in the analysis period |
| Critical Exclusion | Both numerator and denominator exclude responses from queries where brand_in_query = True |
| Rationale | Excluding branded queries prevents inflated mention rates and isolates organic AI platform behavior |

**Example:** If there were 85 total responses from non-branded queries, and your brand was mentioned (directly or indirectly) in 34 of them, the mention rate would be (34/85) × 100 = 40.0%

### Positioning Score

The positioning metric evaluates where your brand appears in AI-generated responses, with higher scores indicating more prominent placement.

**Average Positioning Score Formula:**
```
Average Positioning Score = (Sum of Individual Position Scores) / Total Responses
```

**Position Scoring System:**

| Position Category | Point Value |
|-------------------|-------------|
| Leader | 5 points |
| Top 3 | 4 points |
| Featured | 3 points |
| Listed | 2 points |
| Not Mentioned | 1 point |

Each response receives a score (1-5) based on how your brand was positioned. The scores are summed across all qualifying responses and divided by total response count to produce an average (range: 1.0 to 5.0). Responses from queries where brand_in_query = True are excluded.

**Leadership Visibility (Sub-metric):**
```
Leadership Visibility (%) = ((Leader Count + Top 3 Count) / Total Responses) × 100
```

This metric specifically measures high-quality visibility by combining the top two positioning categories.

### Share of Voice

Share of Voice quantifies your brand's relative visibility compared to all organizations (including competitors) mentioned across AI responses.

**Formula:**
```
Share of Voice (%) = (Brand Mentions / Total All Organization Mentions) × 100
```

| Component | Definition |
|-----------|------------|
| Brand Mentions (Numerator) | Count of responses where your brand achieved positioning of 'Leader', 'Top 3', 'Featured', or 'Listed' |
| Total Mentions (Denominator) | Sum of all organization mentions including: (1) your brand mentions and (2) all competitor mentions |
| Competitor Counting | The competitors field contains comma-separated organization names; each occurrence increments that competitor's mention count |
| Exclusion | Only responses from queries where brand_in_query = False are included |

**Example:** If your brand appeared in 34 responses with qualifying positioning, and competitors appeared in a combined 56 responses, the total mentions would be 90. Your share of voice would be (34/90) × 100 = 37.8%

### Target Descriptor Adoption

Target descriptor adoption measures how successfully your brand has become associated with the specific descriptors and attributes it aims to own strategically.

**Formula:**
```
Descriptor Match Rate (%) = (Number of Target Descriptors Found / Total Target Descriptors) × 100
```

| Component | Definition |
|-----------|------------|
| Total Target Descriptors | Count of all descriptors configured as strategic targets for your brand in the platform |
| Target Descriptors Found | Count of unique target descriptors that appear in at least one AI response |
| Matching Logic | Case-insensitive matching; a target descriptor is counted as "found" if it appears in any response where your brand was mentioned |
| Inclusion | This calculation INCLUDES responses from queries where brand_in_query = True (quality of associations matters regardless of query type) |

**Example:** If your brand has 20 target descriptors and 13 of those descriptors appeared in at least one response, the descriptor match rate would be (13/20) × 100 = 65.0%

### Competitive Threat Analysis

Unlike the quantitative metrics above, competitive threat analysis is **not based on a mathematical formula**. Instead, it employs a qualitative AI-powered methodology to identify strategic competitive risks.

**Process:**

1. **Data Collection Phase:** Gathers Share of Voice data showing competitor mention frequencies, identifies specific query-response pairs where your brand was not mentioned but competitors were, and extracts "competitive loss" examples.

2. **AI Analysis Phase:** Submits concrete response examples and competitive data to Perplexity Sonar Pro model, which identifies patterns in which competitors consistently outperform your brand and analyzes the specific descriptors and positioning competitors have claimed.

3. **Output:** Generates qualitative descriptions of top 3-5 competitive threats, specific examples of queries/responses where competitors won visibility, strategic implications, and recommended counter-actions.

**Rationale:** Competitive threats involve nuanced strategic considerations that resist reduction to single numerical scores. The AI-powered qualitative analysis can weigh multidimensional factors more effectively than a predetermined formula, while remaining data-driven and grounded in actual response examples.

### Summary of Metric Calculation Approaches

| Metric | Type | Includes Branded Queries? | Rationale |
|--------|------|---------------------------|-----------|
| Mention Rate | Quantitative Formula | No (Excluded) | Measures organic visibility without bias |
| Positioning Score | Quantitative Formula | No (Excluded) | Assesses natural positioning |
| Share of Voice | Quantitative Formula | No (Excluded) | Compares competitive visibility organically |
| Descriptor Match | Quantitative Formula | Yes (Included) | Quality of associations matters regardless of query type |
| Sentiment Distribution | Quantitative Formula | Yes (Included) | Sentiment reflects perception across all contexts |
| Competitive Threats | Qualitative AI Analysis | Context-dependent | Strategic nuance requires pattern recognition |

## Competitive Intelligence Analysis

Competitor analysis forms a critical component of the methodology, examining how your brand performs relative to other organizations in AI-generated discourse. The system tracks pre-configured competitors with metadata including organization type, focus areas, and key descriptors, while also automatically extracting mentions of organizations not initially identified as competitors.

Co-occurrence analysis reveals which competitors frequently appear alongside your brand in AI responses, and comparative share of voice calculations quantify relative visibility. The analysis identifies specific queries where competitors received more favorable positioning than your brand, extracts the descriptors and positioning competitors own, and examines concrete response examples showing competitive advantages. These findings enable the generation of targeted recommendations for closing competitive gaps and claiming strategic positioning currently owned by rivals.

## Recommendations Generation

The strategic recommendations produced for your brand result from a sophisticated AI-driven synthesis process that connects quantitative performance data with qualitative strategic context. The system performs gap analysis comparing target descriptors against actual usage, diagnoses which AI platforms underperform and infers why based on response patterns, benchmarks competitive positioning to identify claimable strategic territory, analyzes source types that different platforms prioritize, and incorporates real-time industry developments to identify timely opportunities.

Each recommendation includes strategic rationale backed by specific metrics, four to five tactical action steps with measurable targets, explicit alignment with target descriptors your brand aims to reinforce, source strategies specifying what content to create and where to publish it, and platform targeting identifying which AI systems each tactic aims to influence. This approach ensures recommendations are data-driven, actionable, and directly tied to measurable improvements in AI reputation metrics.

## Limitations and Considerations

Several important limitations should be considered when interpreting these findings. AI platform responses can vary over time due to model updates, training data changes, and index refreshes, meaning that findings represent a snapshot rather than static characteristics. Finally, while the analysis identifies correlations between content strategies and AI reputation metrics, establishing direct causation requires controlled longitudinal studies that account for confounding variables such as broader industry trends, news coverage, and publication timing effects.
"""

    # Generate Word document
    word_file = export_to_word(methodology_content, "How Tales Works")

    return StreamingResponse(
        word_file,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": "attachment; filename=How_Tales_Works.docx"}
    )


@app.get("/reports/{report_id}/export/html", tags=["Reports"])
def export_report_to_html(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Export a report to interactive HTML with embedded Chart.js visualizations."""
    from app.services.report_html import generate_html_report_with_charts

    db_report = crud.get_report(db, report_id=report_id, user_id=current_user.id)
    if db_report is None:
        raise HTTPException(status_code=404, detail="Report not found")

    # Generate HTML with charts
    html_content = generate_html_report_with_charts(
        db_report.report_content,
        db_report.title,
        db,
        user_id=current_user.id,
        brand_id=brand_id
    )

    # Create safe filename
    safe_filename = "".join(c for c in db_report.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"{safe_filename}.html"

    return StreamingResponse(
        io.BytesIO(html_content.encode('utf-8')),
        media_type="text/html",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# --- Brand Info Endpoints ---
# --- Multi-Brand Endpoints ---

@app.get("/brands/", response_model=List[schemas.BrandInfoList], tags=["Brands"])
def get_all_brands_endpoint(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all brands for the current user."""
    return crud.get_all_brands(db, user_id=current_user.id)

@app.get("/brands/active", response_model=schemas.BrandInfo, tags=["Brands"])
def get_active_brand_endpoint(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get the currently active brand for the user."""
    brand = crud.get_active_brand(db, user_id=current_user.id)
    if not brand:
        raise HTTPException(status_code=404, detail="No active brand found. Please create a brand first.")
    return brand

@app.get("/brands/{brand_id}", response_model=schemas.BrandInfo, tags=["Brands"])
def get_brand_endpoint(
    brand_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific brand by ID."""
    brand = crud.get_brand_by_id(db, brand_id=brand_id, user_id=current_user.id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    return brand

@app.post("/brands/", response_model=schemas.BrandInfo, status_code=201, tags=["Brands"])
def create_brand_endpoint(
    brand_info: schemas.BrandInfoCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new brand (max 20 per user)."""
    try:
        return crud.create_brand_info(db, brand_info=brand_info, user_id=current_user.id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/brands/{brand_id}", response_model=schemas.BrandInfo, tags=["Brands"])
def update_brand_endpoint(
    brand_id: int,
    brand_info: schemas.BrandInfoUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a specific brand."""
    updated_brand = crud.update_brand_info(db, brand_id=brand_id, brand_info_update=brand_info, user_id=current_user.id)
    if not updated_brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    return updated_brand

@app.post("/brands/{brand_id}/activate", response_model=schemas.BrandInfo, tags=["Brands"])
def activate_brand_endpoint(
    brand_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Set a brand as active (deactivates all other brands)."""
    brand = crud.activate_brand(db, brand_id=brand_id, user_id=current_user.id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    return brand

@app.delete("/brands/{brand_id}", response_model=schemas.BrandInfo, tags=["Brands"])
def delete_brand_endpoint(
    brand_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a brand and all associated data."""
    deleted_brand = crud.delete_brand_info(db, brand_id=brand_id, user_id=current_user.id)
    if not deleted_brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    return deleted_brand

@app.post("/brands/{brand_id}/share", response_model=schemas.BrandShare, status_code=201, tags=["Brands"])
def share_brand_endpoint(
    brand_id: int,
    share_data: schemas.BrandShareCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Share a brand with another user by email."""
    # Verify the current user has access to this brand (owns it or has it shared)
    if not crud.user_has_brand_access(db, brand_id, current_user.id):
        raise HTTPException(status_code=404, detail="Brand not found or you don't have permission to share it")

    brand = db.query(models.BrandInfo).filter(
        models.BrandInfo.id == brand_id
    ).first()

    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    # Find the user to share with by email
    user_to_share_with = db.query(models.User).filter(models.User.email == share_data.email).first()

    if not user_to_share_with:
        raise HTTPException(status_code=404, detail=f"No user found with email {share_data.email}. User must have a TALES account.")

    if user_to_share_with.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot share a brand with yourself")

    # Check if already shared with this user
    existing_share = db.query(models.BrandShare).filter(
        models.BrandShare.brand_id == brand_id,
        models.BrandShare.user_id == user_to_share_with.id
    ).first()

    if existing_share:
        raise HTTPException(status_code=400, detail=f"Brand is already shared with {share_data.email}")

    # Create the share
    new_share = models.BrandShare(
        brand_id=brand_id,
        user_id=user_to_share_with.id,
        shared_by_user_id=current_user.id,
        permission_level='edit'
    )
    db.add(new_share)
    db.commit()
    db.refresh(new_share)

    return new_share

@app.get("/brands/{brand_id}/shares", response_model=List[schemas.BrandShareWithUser], tags=["Brands"])
def get_brand_shares_endpoint(
    brand_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all users this brand is shared with."""
    # Verify user has access to this brand
    if not crud.user_has_brand_access(db, brand_id, current_user.id):
        raise HTTPException(status_code=404, detail="Brand not found or you don't have permission to view shares")

    # Get all shares with user details
    shares = db.query(models.BrandShare).filter(
        models.BrandShare.brand_id == brand_id
    ).all()

    # Build response with user details
    result = []
    for share in shares:
        user = db.query(models.User).filter(models.User.id == share.user_id).first()
        shared_by = db.query(models.User).filter(models.User.id == share.shared_by_user_id).first()

        result.append(schemas.BrandShareWithUser(
            id=share.id,
            brand_id=share.brand_id,
            user_id=share.user_id,
            user_email=user.email if user else "Unknown",
            user_full_name=user.full_name if user else None,
            shared_by_user_id=share.shared_by_user_id,
            shared_by_email=shared_by.email if shared_by else "Unknown",
            permission_level=share.permission_level,
            created_at=share.created_at
        ))

    return result

@app.delete("/brands/{brand_id}/shares/{share_id}", tags=["Brands"])
def remove_brand_share_endpoint(
    brand_id: int,
    share_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Remove a brand share (unshare with a user). Only the brand owner can remove shares."""
    # Verify user OWNS this brand (not just has access)
    brand = db.query(models.BrandInfo).filter(
        models.BrandInfo.id == brand_id,
        models.BrandInfo.user_id == current_user.id
    ).first()

    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found or only the brand owner can remove shares")

    # Find and delete the share
    share = db.query(models.BrandShare).filter(
        models.BrandShare.id == share_id,
        models.BrandShare.brand_id == brand_id
    ).first()

    if not share:
        raise HTTPException(status_code=404, detail="Share not found")

    db.delete(share)
    db.commit()

    return {"message": "Share removed successfully"}

# --- Legacy Endpoints for Backward Compatibility ---

@app.get("/brand-info/", response_model=schemas.BrandInfo, tags=["Brand Info"])
def get_brand_info_endpoint(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get active brand info (legacy endpoint - use /brands/active instead)."""
    brand_info = crud.get_active_brand(db, user_id=current_user.id)
    if not brand_info:
        raise HTTPException(status_code=404, detail="Brand info not found. Please create it first.")
    return brand_info

@app.post("/brand-info/", response_model=schemas.BrandInfo, status_code=201, tags=["Brand Info"])
def create_brand_info_endpoint(
    brand_info: schemas.BrandInfoCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create brand info (legacy endpoint - use /brands/ instead)."""
    try:
        return crud.create_brand_info(db, brand_info=brand_info, user_id=current_user.id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/brand-info/", response_model=schemas.BrandInfo, tags=["Brand Info"])
def update_brand_info_endpoint(
    brand_info: schemas.BrandInfoUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update active brand info (legacy endpoint)."""
    active_brand = crud.get_active_brand(db, user_id=current_user.id)
    if not active_brand:
        raise HTTPException(status_code=404, detail="No active brand found")
    updated_brand = crud.update_brand_info(db, brand_id=active_brand.id, brand_info_update=brand_info, user_id=current_user.id)
    if not updated_brand:
        raise HTTPException(status_code=404, detail="Brand info not found")
    return updated_brand

@app.delete("/brand-info/", response_model=schemas.BrandInfo, tags=["Brand Info"])
def delete_brand_info_endpoint(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete active brand info (legacy endpoint)."""
    active_brand = crud.get_active_brand(db, user_id=current_user.id)
    if not active_brand:
        raise HTTPException(status_code=404, detail="No active brand found")
    deleted_brand = crud.delete_brand_info(db, brand_id=active_brand.id, user_id=current_user.id)
    if not deleted_brand:
        raise HTTPException(status_code=404, detail="Brand info not found")
    return deleted_brand


@app.post("/brand-info/generate", tags=["Brand Info"])
def generate_content_from_brand_info(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """
    Generate queries, descriptors, and competitors based on active brand info using AI.
    This will replace any existing queries, descriptors, and competitors for the active brand.
    """
    from app.ai_generator import AIGenerator

    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand found. Please select a brand first.")

    try:
        generator = AIGenerator(db)
        result = generator.generate_all(user_id=current_user.id, brand_id=brand_id)
        return {
            "message": "Content generated successfully",
            "queries_created": result["queries_created"],
            "descriptors_created": result["descriptors_created"],
            "competitors_created": result["competitors_created"]
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate content: {str(e)}")


# --- Celery Task Trigger for the main weekly run ---
from celery_app.tasks import run_weekly_queries_task
@app.post("/tasks/trigger-weekly-run/", status_code=202, tags=["Tasks"])
async def trigger_weekly_run_endpoint():
    """Manually trigger the weekly query and analysis process."""
    task = run_weekly_queries_task.delay()
    return {"message": "Weekly run triggered.", "task_id": task.id}

@app.post("/tasks/trigger-analysis/", status_code=202, tags=["Tasks"])
async def trigger_analysis_on_unanalyzed(db: Session = Depends(get_db)):
    """Manually trigger analysis on all unanalyzed responses."""
    from celery_app.tasks import analyze_responses_batch_task
    unanalyzed_responses = crud.get_unanalyzed_responses(db, limit=1000)
    if not unanalyzed_responses:
        raise HTTPException(status_code=404, detail="No unanalyzed responses found.")

    response_ids = [resp.id for resp in unanalyzed_responses]
    task = analyze_responses_batch_task.delay(response_ids)
    return {"message": f"Analysis triggered for {len(response_ids)} responses.", "task_id": task.id}

# --- Direct Collection and Analysis Triggers (without Celery) ---
@app.post("/tasks/run-collection/", status_code=202, tags=["Tasks"])
async def run_collection(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Trigger response collection using the collection script for active brand."""
    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand found. Please select a brand first.")

    # Get count of active queries
    query_count = db.query(models.Query).filter(
        models.Query.user_id == current_user.id,
        models.Query.brand_id == brand_id,
        models.Query.active == True
    ).count()

    if query_count == 0:
        raise HTTPException(status_code=400, detail="No active queries found for this brand.")

    # Create task status for collection
    task_status = models.TaskStatus(
        user_id=current_user.id,
        brand_id=brand_id,
        task_type="collection",
        status="running",
        total_items=query_count * 4,  # queries × 4 platforms
        processed_items=0,
        message="Starting collection..."
    )
    db.add(task_status)
    db.commit()
    db.refresh(task_status)

    script_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "collect_responses.py")
    try:
        # Run the collection script in the background with task-id
        cmd = [
            "python3", script_path,
            str(current_user.id),
            "--brand-id", str(brand_id),
            "--task-id", str(task_status.id)
        ]
        process = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        # Send "1" to run all queries (auto-select option 1)
        process.stdin.write("1\n")
        process.stdin.flush()

        return {
            "message": "Response collection started for active brand.",
            "status": "running",
            "task_id": task_status.id,
            "note": "Collection is running in the background. Check the Responses page to see new data."
        }
    except Exception as e:
        task_status.status = "failed"
        task_status.error_message = str(e)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to start collection: {str(e)}")

@app.post("/tasks/run-analysis/", status_code=202, tags=["Tasks"])
async def run_analysis(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Analyze responses from most recent collection date and auto-generate report for active brand."""
    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand found. Please select a brand first.")

    # Find the most recent collection date for this brand
    most_recent_response = db.query(models.Response).filter(
        models.Response.user_id == current_user.id,
        models.Response.brand_id == brand_id
    ).order_by(models.Response.timestamp.desc()).first()

    if not most_recent_response:
        raise HTTPException(status_code=404, detail="No responses found for active brand.")

    # Get the date of the most recent collection (just the date part)
    latest_date = most_recent_response.timestamp.date()

    # Find all responses from that date
    latest_date_start = datetime.datetime.combine(latest_date, datetime.time.min)
    latest_date_end = latest_date_start + datetime.timedelta(days=1)

    responses_to_analyze = db.query(models.Response).filter(
        models.Response.user_id == current_user.id,
        models.Response.brand_id == brand_id,
        models.Response.timestamp >= latest_date_start,
        models.Response.timestamp < latest_date_end
    ).all()

    if not responses_to_analyze:
        raise HTTPException(status_code=404, detail="No responses found for the most recent collection date.")

    # Reset analysis fields for these responses
    for response in responses_to_analyze:
        response.analyzed_at = None
        response.brand_mentioned = None
        response.brand_position = None
        response.sentiment = None
        response.descriptors = None
        response.competitors = None
        response.sources = None

    db.commit()

    # Create task status for analysis
    task_status = models.TaskStatus(
        user_id=current_user.id,
        brand_id=brand_id,
        task_type="analysis_and_report",
        status="running",
        total_items=len(responses_to_analyze),
        processed_items=0,
        message=f"Analyzing {len(responses_to_analyze)} responses from {latest_date.strftime('%Y-%m-%d')}..."
    )
    db.add(task_status)
    db.commit()
    db.refresh(task_status)

    analysis_script = os.path.join(os.path.dirname(os.path.dirname(__file__)), "analyze_responses.py")
    report_script = os.path.join(os.path.dirname(os.path.dirname(__file__)), "generate_report.py")

    try:
        # Get the project root directory (where tales.db is located)
        project_root = os.path.dirname(os.path.dirname(__file__))

        # Run analysis and report generation in sequence in the background with task-id
        # Use separate commands to capture which step fails
        analysis_cmd = f"python3 {analysis_script} --all --user-id {current_user.id} --brand-id {brand_id} --task-id {task_status.id}"
        report_cmd = f"python3 {report_script} --user-id {current_user.id} --brand-id {brand_id}"

        # Create a background task to monitor the subprocess
        import threading

        def run_analysis_task():
            """Run analysis and report generation, updating task status on completion."""
            db_task = SessionLocal()
            try:
                # Run analysis script
                analysis_process = subprocess.run(
                    analysis_cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    cwd=project_root,
                    timeout=3600  # 1 hour timeout
                )

                if analysis_process.returncode != 0:
                    # Analysis failed
                    task = db_task.query(models.TaskStatus).filter(
                        models.TaskStatus.id == task_status.id
                    ).first()
                    if task:
                        task.status = "failed"
                        task.error_message = f"Analysis script failed: {analysis_process.stderr[-500:]}"
                        db_task.commit()
                    return

                # Run report generation script
                report_process = subprocess.run(
                    report_cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    cwd=project_root,
                    timeout=600  # 10 minute timeout
                )

                if report_process.returncode != 0:
                    # Report generation failed
                    task = db_task.query(models.TaskStatus).filter(
                        models.TaskStatus.id == task_status.id
                    ).first()
                    if task:
                        task.status = "failed"
                        task.error_message = f"Report generation failed: {report_process.stderr[-500:]}"
                        db_task.commit()
                    return

                # Both succeeded
                task = db_task.query(models.TaskStatus).filter(
                    models.TaskStatus.id == task_status.id
                ).first()
                if task:
                    task.status = "completed"
                    task.message = "Analysis and report generation completed successfully"
                    db_task.commit()

            except subprocess.TimeoutExpired as e:
                task = db_task.query(models.TaskStatus).filter(
                    models.TaskStatus.id == task_status.id
                ).first()
                if task:
                    task.status = "failed"
                    task.error_message = f"Task timed out: {str(e)}"
                    db_task.commit()
            except Exception as e:
                task = db_task.query(models.TaskStatus).filter(
                    models.TaskStatus.id == task_status.id
                ).first()
                if task:
                    task.status = "failed"
                    task.error_message = f"Unexpected error: {str(e)}"
                    db_task.commit()
            finally:
                db_task.close()

        # Start the background thread
        thread = threading.Thread(target=run_analysis_task, daemon=True)
        thread.start()

        return {
            "message": f"Analysis and report generation started for {len(responses_to_analyze)} responses from {latest_date.strftime('%Y-%m-%d')}.",
            "status": "running",
            "task_id": task_status.id,
            "count": len(responses_to_analyze),
            "note": "Analyzing latest data and generating report. This will update all analytics pages."
        }
    except Exception as e:
        task_status.status = "failed"
        task_status.error_message = str(e)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to start analysis: {str(e)}")

@app.post("/tasks/rerun-analysis/", status_code=202, tags=["Tasks"])
async def rerun_analysis(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Reset analysis on responses for active brand (optionally filtered by date range) and regenerate report."""
    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand found. Please select a brand first.")

    # Build query for responses
    query = db.query(models.Response).filter(
        models.Response.user_id == current_user.id,
        models.Response.brand_id == brand_id
    )

    # Apply date filters if provided
    date_description = "all"
    if start_date:
        try:
            start_dt = datetime.datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(models.Response.timestamp >= start_dt)
            date_description = f"from {start_date[:10]}"
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use ISO format (YYYY-MM-DD).")

    if end_date:
        try:
            end_dt = datetime.datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            # Add one day to include the entire end date
            end_dt = end_dt + datetime.timedelta(days=1)
            query = query.filter(models.Response.timestamp < end_dt)
            if start_date:
                date_description = f"from {start_date[:10]} to {end_date[:10]}"
            else:
                date_description = f"through {end_date[:10]}"
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use ISO format (YYYY-MM-DD).")

    all_responses = query.all()

    if not all_responses:
        raise HTTPException(status_code=404, detail=f"No responses found for active brand {date_description}.")

    # Reset analyzed_at to NULL for filtered responses (marks them as unanalyzed)
    for response in all_responses:
        response.analyzed_at = None
        # Clear existing analysis fields so they get fresh analysis
        response.brand_mentioned = None
        response.brand_position = None
        response.sentiment = None
        response.descriptors = None
        response.competitors = None
        response.sources = None

    db.commit()

    # Create task status for re-analysis
    message = f"Re-analyzing {len(all_responses)} responses {date_description}..."
    task_status = models.TaskStatus(
        user_id=current_user.id,
        brand_id=brand_id,
        task_type="analysis_and_report",
        status="running",
        total_items=len(all_responses),
        processed_items=0,
        message=message,
        started_at=datetime.datetime.utcnow()
    )
    db.add(task_status)
    db.commit()
    db.refresh(task_status)

    analysis_script = os.path.join(os.path.dirname(os.path.dirname(__file__)), "analyze_responses.py")
    report_script = os.path.join(os.path.dirname(os.path.dirname(__file__)), "generate_report.py")

    try:
        # Get the project root directory (where tales.db is located)
        project_root = os.path.dirname(os.path.dirname(__file__))

        # Run analysis and report generation in sequence in the background
        # Use separate commands to capture which step fails
        analysis_cmd = f"python3 {analysis_script} --all --user-id {current_user.id} --brand-id {brand_id} --task-id {task_status.id}"
        report_cmd = f"python3 {report_script} --user-id {current_user.id} --brand-id {brand_id}"

        # Create a background task to monitor the subprocess
        import threading

        def run_reanalysis_task():
            """Run re-analysis and report generation, updating task status on completion."""
            db_task = SessionLocal()
            try:
                # Run analysis script
                analysis_process = subprocess.run(
                    analysis_cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    cwd=project_root,
                    timeout=3600  # 1 hour timeout
                )

                if analysis_process.returncode != 0:
                    # Analysis failed
                    task = db_task.query(models.TaskStatus).filter(
                        models.TaskStatus.id == task_status.id
                    ).first()
                    if task:
                        task.status = "failed"
                        task.error_message = f"Analysis script failed: {analysis_process.stderr[-500:]}"
                        db_task.commit()
                    return

                # Run report generation script
                report_process = subprocess.run(
                    report_cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    cwd=project_root,
                    timeout=600  # 10 minute timeout
                )

                if report_process.returncode != 0:
                    # Report generation failed
                    task = db_task.query(models.TaskStatus).filter(
                        models.TaskStatus.id == task_status.id
                    ).first()
                    if task:
                        task.status = "failed"
                        task.error_message = f"Report generation failed: {report_process.stderr[-500:]}"
                        db_task.commit()
                    return

                # Both succeeded
                task = db_task.query(models.TaskStatus).filter(
                    models.TaskStatus.id == task_status.id
                ).first()
                if task:
                    task.status = "completed"
                    task.message = "Re-analysis and report generation completed successfully"
                    db_task.commit()

            except subprocess.TimeoutExpired as e:
                task = db_task.query(models.TaskStatus).filter(
                    models.TaskStatus.id == task_status.id
                ).first()
                if task:
                    task.status = "failed"
                    task.error_message = f"Task timed out: {str(e)}"
                    db_task.commit()
            except Exception as e:
                task = db_task.query(models.TaskStatus).filter(
                    models.TaskStatus.id == task_status.id
                ).first()
                if task:
                    task.status = "failed"
                    task.error_message = f"Unexpected error: {str(e)}"
                    db_task.commit()
            finally:
                db_task.close()

        # Start the background thread
        thread = threading.Thread(target=run_reanalysis_task, daemon=True)
        thread.start()

        return {
            "message": f"Re-analysis started for {len(all_responses)} responses {date_description}.",
            "status": "running",
            "task_id": task_status.id,
            "count": len(all_responses),
            "date_range": date_description,
            "note": "Re-analyzing responses with updated analysis process. A new report will be generated when complete."
        }
    except Exception as e:
        task_status.status = "failed"
        task_status.error_message = str(e)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to start re-analysis: {str(e)}")

@app.get("/tasks/status/", response_model=Optional[schemas.TaskStatus], tags=["Tasks"])
def get_task_status(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Get the status of the most recent task for the active brand."""
    if not brand_id:
        return None

    # Get the most recent task for this user and brand
    task = db.query(models.TaskStatus).filter(
        models.TaskStatus.user_id == current_user.id,
        models.TaskStatus.brand_id == brand_id
    ).order_by(models.TaskStatus.started_at.desc()).first()

    if not task:
        return None

    # Check if the task is still running by checking unanalyzed responses
    if task.status == "running":
        unanalyzed = crud.get_unanalyzed_responses(db, user_id=current_user.id, brand_id=brand_id, limit=1)
        if not unanalyzed and task.task_type == "analysis_and_report":
            # Analysis is complete, check if we're generating report
            # For simplicity, mark as completed if no unanalyzed responses
            task.status = "completed"
            task.completed_at = datetime.datetime.utcnow()
            task.message = "Analysis and report generation completed"
            db.commit()
            db.refresh(task)

    return task
