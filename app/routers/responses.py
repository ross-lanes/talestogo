"""
Responses API Endpoints
Provides CRUD operations, bulk operations, and export functionality for responses
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io

from .. import crud, models, schemas
from ..auth import get_current_user
from ..database import SessionLocal, get_db
from ..utils.brand_access import get_active_brand_id, get_data_owner_user_id


router = APIRouter(
    prefix="/responses",
    tags=["Responses"]
)


@router.post("/", response_model=schemas.Response, status_code=201)
def create_response(
    response: schemas.ResponseCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Submit a raw response from an LLM platform for the current user's active brand."""
    return crud.create_response(db=db, response=response, user_id=current_user.id, brand_id=brand_id)


@router.get("/", response_model=List[schemas.Response])
def read_responses(
    skip: int = 0,
    limit: int = 10000,
    batch_id: Optional[int] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id),
):
    """Retrieve responses for the active brand.

    Resolves the brand owner's user_id so that users viewing a brand that's
    been shared with them see the owner's data (Bug #1 fix). Accepts a
    `batch_id` query param to narrow to a single collection batch (Bug #2 fix).
    Default limit raised to 10000 to cover real-world brand sizes like PPPL
    (~1,200 responses); matches the precedent used by the Excel export below
    (Bug #3 fix).

    `limit` is defensively clamped to [1, 10000] so a malicious or buggy
    client can't request a negative or oversized page that would either
    return nothing or pull the entire table into memory.
    """
    limit = max(1, min(limit, 10000))
    owner_user_id = get_data_owner_user_id(db, brand_id, current_user.id)
    return crud.get_responses(
        db,
        user_id=owner_user_id,
        brand_id=brand_id,
        batch_id=batch_id,
        skip=skip,
        limit=limit,
    )


@router.get("/unanalyzed/", response_model=List[schemas.Response])
def read_unanalyzed_responses(
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Retrieve responses that are pending analysis for the current user."""
    return crud.get_unanalyzed_responses(db, user_id=current_user.id, limit=limit)


@router.put("/{response_id}/analyze", response_model=schemas.Response)
def update_response_analysis(
    response_id: int,
    analysis_data: schemas.ResponseAnalysisInput,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a response with analysis data for the current user's active brand."""
    db_response = crud.update_response_analysis(
        db,
        response_id=response_id,
        analysis_data=analysis_data.model_dump(exclude_unset=True),
        user_id=current_user.id,
        brand_id=brand_id
    )
    if db_response is None:
        raise HTTPException(status_code=404, detail="Response not found")
    return db_response


@router.delete("/{response_id}", response_model=schemas.Response)
def delete_response(
    response_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a response for the current user's active brand."""
    deleted_response = crud.delete_response(db, response_id=response_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_response is None:
        raise HTTPException(status_code=404, detail="Response not found")
    return deleted_response


@router.post("/bulk-replace-competitor")
def bulk_replace_competitor(
    old_name: str,
    new_name: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """
    Bulk find-and-replace competitor names across all responses for the active brand.

    Args:
        old_name: The competitor name to find (exact match)
        new_name: The new name to replace it with

    Returns:
        Number of responses updated
    """
    # Get all responses for the active brand
    responses = crud.get_responses(db, user_id=current_user.id, brand_id=brand_id, skip=0, limit=100000)

    updated_count = 0

    for response in responses:
        if response.competitors:
            # Split by comma, find and replace exact matches, rejoin
            competitors_list = [comp.strip() for comp in response.competitors.split(',')]

            # Replace exact matches (case-sensitive)
            updated_list = [new_name if comp == old_name else comp for comp in competitors_list]

            new_competitors = ', '.join(updated_list)

            if new_competitors != response.competitors:
                response.competitors = new_competitors
                updated_count += 1

    db.commit()

    return {"updated_count": updated_count, "old_name": old_name, "new_name": new_name}


@router.get("/export/excel")
def export_responses_to_excel(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Export all responses for the active brand to an Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    # Get all responses for the active brand
    responses = crud.get_responses(db, user_id=current_user.id, brand_id=brand_id, skip=0, limit=10000)

    if not responses:
        raise HTTPException(status_code=404, detail="No responses found for export")

    # Get brand name for filename
    brand = crud.get_active_brand(db, user_id=current_user.id) if brand_id else None
    brand_name = brand.brand_name if brand else "Unknown"

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Responses"

    # Define headers
    headers = [
        "Response ID",
        "Query ID",
        "Platform",
        "Response Text",
        "Collected At",
        "Analyzed At",
        "Brand Mentioned",
        "Brand Position",
        "Sentiment",
        "Descriptors",
        "Competitors",
        "Sources",
        "Notes"
    ]

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for row_num, response in enumerate(responses, 2):
        ws.cell(row=row_num, column=1).value = response.id
        ws.cell(row=row_num, column=2).value = response.query_id
        ws.cell(row=row_num, column=3).value = response.platform
        ws.cell(row=row_num, column=4).value = response.response_text
        ws.cell(row=row_num, column=5).value = response.timestamp.strftime('%Y-%m-%d %H:%M:%S') if response.timestamp else ''
        ws.cell(row=row_num, column=6).value = response.analyzed_at.strftime('%Y-%m-%d %H:%M:%S') if response.analyzed_at else ''
        ws.cell(row=row_num, column=7).value = response.brand_mentioned or ''
        ws.cell(row=row_num, column=8).value = response.brand_position or ''
        ws.cell(row=row_num, column=9).value = response.sentiment or ''
        ws.cell(row=row_num, column=10).value = response.descriptors or ''
        ws.cell(row=row_num, column=11).value = response.competitors or ''
        ws.cell(row=row_num, column=12).value = response.sources or ''
        ws.cell(row=row_num, column=13).value = response.notes or ''

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create safe filename
    safe_brand_name = "".join(c for c in brand_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"{safe_brand_name}_AI_Responses.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
