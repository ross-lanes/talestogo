"""
Query API Endpoints
CRUD operations for queries and bulk upload functionality
"""
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.orm import Session
from typing import List, Optional
import io
from openpyxl import load_workbook

from .. import crud, models, schemas
from ..auth import get_current_user, get_current_admin_user
from ..database import get_db
from ..routers.analytics import get_active_brand_id
from ..utils.file_validation import validate_excel_upload_with_size_limit

router = APIRouter(
    prefix="/queries",
    tags=["Queries"]
)


@router.post("/", response_model=schemas.Query, status_code=201)
def create_query(
    query: schemas.QueryCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Create a new query for the current user's active brand."""
    db_query = crud.get_query_by_query_id(db, query_id=query.query_id, user_id=current_user.id, brand_id=brand_id)
    if db_query:
        raise HTTPException(status_code=400, detail=f"Query ID {query.query_id} already exists for this brand")
    return crud.create_query(db=db, query=query, user_id=current_user.id, brand_id=brand_id)


@router.get("/", response_model=List[schemas.Query])
def read_queries(
    active_only: bool = False,
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of queries for the current user's active brand."""
    if active_only:
        queries = crud.get_active_queries(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)
    else:
        queries = crud.get_queries(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)
    return queries


@router.get("/{query_id}", response_model=schemas.Query)
def read_query(
    query_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a single query by its user-facing ID (e.g., 'Q001') for the current user's active brand."""
    db_query = crud.get_query_by_query_id(db, query_id=query_id, user_id=current_user.id, brand_id=brand_id)
    if db_query is None:
        raise HTTPException(status_code=404, detail="Query not found")
    return db_query


@router.put("/{query_id}", response_model=schemas.Query)
def update_query(
    query_id: str,
    query_update: schemas.QueryUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a query for the current user's active brand."""
    db_query = crud.update_query(db, query_id=query_id, query_update=query_update, user_id=current_user.id, brand_id=brand_id)
    if db_query is None:
        raise HTTPException(status_code=404, detail="Query not found")
    return db_query


@router.delete("/{query_id}", response_model=schemas.Query)
def delete_query(
    query_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a query for the current user's active brand."""
    deleted_query = crud.delete_query(db, query_id=query_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_query is None:
        raise HTTPException(status_code=404, detail="Query not found")
    return deleted_query


@router.post("/upload")
async def upload_queries(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Upload queries from an Excel file (admin only). Expected columns: query_id, query_text, category, target_outcome, brand_in_query, active"""

    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand selected")

    # Get the brand to determine the owner's user_id
    brand = db.query(models.BrandInfo).filter(models.BrandInfo.id == brand_id).first()
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    # Use the brand owner's user_id, not the admin's user_id
    # This ensures data ownership remains with the brand owner
    owner_user_id = brand.user_id

    try:
        # Validate file type and size using magic byte checking
        contents = await validate_excel_upload_with_size_limit(file)

        # Read the Excel file
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Validate required columns
        required_columns = ['query_id', 'query_text']
        for col in required_columns:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        created_count = 0
        updated_count = 0
        errors = []

        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dict from row
                row_data = dict(zip(headers, row))

                if not row_data.get('query_id'):
                    continue  # Skip empty rows

                # Check if query already exists
                existing_query = db.query(models.Query).filter(
                    models.Query.query_id == row_data['query_id'],
                    models.Query.user_id == owner_user_id,
                    models.Query.brand_id == brand_id
                ).first()

                if existing_query:
                    # Update existing query
                    existing_query.query_text = row_data.get('query_text', existing_query.query_text)
                    existing_query.category = row_data.get('category', existing_query.category)
                    existing_query.target_outcome = row_data.get('target_outcome', existing_query.target_outcome)
                    existing_query.brand_in_query = row_data.get('brand_in_query', existing_query.brand_in_query)
                    if 'active' in row_data and row_data['active'] is not None:
                        existing_query.active = bool(row_data['active'])
                    updated_count += 1
                else:
                    # Create new query
                    new_query = models.Query(
                        user_id=owner_user_id,  # Use brand owner's ID, not admin's ID
                        brand_id=brand_id,
                        query_id=row_data['query_id'],
                        query_text=row_data['query_text'],
                        category=row_data.get('category'),
                        target_outcome=row_data.get('target_outcome'),
                        brand_in_query=row_data.get('brand_in_query', False),
                        active=row_data.get('active', True) if row_data.get('active') is not None else True
                    )
                    db.add(new_query)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        db.commit()

        return {
            "message": "Queries uploaded successfully",
            "created": created_count,
            "updated": updated_count,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
