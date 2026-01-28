"""
Descriptor CRUD and upload endpoints.
Provides endpoints for managing target descriptors with Excel upload functionality.
"""
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.orm import Session
from typing import List, Optional
import io

from .. import crud, models, schemas
from ..auth import get_current_user, get_current_admin_user
from ..database import get_db
from ..routers.analytics import get_active_brand_id
from ..utils.file_validation import validate_excel_upload_with_size_limit

router = APIRouter(
    prefix="/descriptors",
    tags=["Descriptors"]
)


@router.post("/", response_model=schemas.TargetDescriptor, status_code=201)
def create_descriptor(
    descriptor: schemas.TargetDescriptorCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Create a new descriptor for the current user's active brand."""
    return crud.create_descriptor(db=db, descriptor=descriptor, user_id=current_user.id, brand_id=brand_id)


@router.get("/", response_model=List[schemas.TargetDescriptor])
def read_descriptors(
    target_only: bool = False,
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of descriptors for the current user's active brand."""
    if target_only:
        return crud.get_target_descriptors(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)
    return crud.get_descriptors(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)


@router.get("/{descriptor_id}", response_model=schemas.TargetDescriptor)
def read_descriptor(
    descriptor_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a single descriptor by its primary key for the current user's active brand."""
    db_descriptor = crud.get_descriptor(db, descriptor_id=descriptor_id, user_id=current_user.id, brand_id=brand_id)
    if db_descriptor is None:
        raise HTTPException(status_code=404, detail="Descriptor not found")
    return db_descriptor


@router.put("/{descriptor_id}", response_model=schemas.TargetDescriptor)
def update_descriptor(
    descriptor_id: int,
    descriptor_update: schemas.TargetDescriptorUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a descriptor for the current user's active brand."""
    db_descriptor = crud.update_descriptor(
        db,
        descriptor_id=descriptor_id,
        descriptor_update=descriptor_update,
        user_id=current_user.id,
        brand_id=brand_id
    )
    if db_descriptor is None:
        raise HTTPException(status_code=404, detail="Descriptor not found")
    return db_descriptor


@router.delete("/{descriptor_id}", response_model=schemas.TargetDescriptor)
def delete_descriptor(
    descriptor_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a descriptor for the current user's active brand."""
    deleted_descriptor = crud.delete_descriptor(db, descriptor_id=descriptor_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_descriptor is None:
        raise HTTPException(status_code=404, detail="Descriptor not found")
    return deleted_descriptor


@router.post("/upload")
async def upload_descriptors(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Upload descriptors from an Excel file (admin only). Expected columns: descriptor, is_target, current_ownership, priority, notes"""
    from openpyxl import load_workbook

    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand selected")

    # Get the brand to determine the owner's user_id
    brand = db.query(models.BrandInfo).filter(models.BrandInfo.id == brand_id).first()
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    # Use the brand owner's user_id, not the admin's user_id
    # This ensures data ownership remains with the brand owner
    owner_user_id = brand.user_id

    try:
        # Validate file type and size using magic byte checking
        contents = await validate_excel_upload_with_size_limit(file)

        # Read the Excel file
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Validate required columns
        required_columns = ['descriptor']
        for col in required_columns:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        created_count = 0
        updated_count = 0
        errors = []

        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dict from row
                row_data = dict(zip(headers, row))

                if not row_data.get('descriptor'):
                    continue  # Skip empty rows

                # Check if descriptor already exists
                existing_descriptor = db.query(models.TargetDescriptor).filter(
                    models.TargetDescriptor.descriptor == row_data['descriptor'],
                    models.TargetDescriptor.user_id == owner_user_id,
                    models.TargetDescriptor.brand_id == brand_id
                ).first()

                if existing_descriptor:
                    # Update existing descriptor
                    if 'is_target' in row_data and row_data['is_target'] is not None:
                        existing_descriptor.is_target = bool(row_data['is_target'])
                    if 'current_ownership' in row_data:
                        existing_descriptor.current_ownership = row_data.get('current_ownership')
                    if 'priority' in row_data:
                        existing_descriptor.priority = row_data.get('priority')
                    if 'notes' in row_data:
                        existing_descriptor.notes = row_data.get('notes')
                    updated_count += 1
                else:
                    # Create new descriptor
                    new_descriptor = models.TargetDescriptor(
                        user_id=owner_user_id,  # Use brand owner's ID, not admin's ID
                        brand_id=brand_id,
                        descriptor=row_data['descriptor'],
                        is_target=row_data.get('is_target', True) if row_data.get('is_target') is not None else True,
                        current_ownership=row_data.get('current_ownership'),
                        priority=row_data.get('priority', 'Medium'),
                        notes=row_data.get('notes')
                    )
                    db.add(new_descriptor)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        db.commit()

        return {
            "message": "Descriptors uploaded successfully",
            "created": created_count,
            "updated": updated_count,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
