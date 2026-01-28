"""
Competitors API Endpoints
CRUD and upload endpoints for managing competitors
"""
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.orm import Session
from typing import List, Optional
import io
from openpyxl import load_workbook

from .. import crud, models, schemas
from ..auth import get_current_user, get_current_admin_user
from ..database import get_db
from ..routers.analytics import get_active_brand_id
from ..utils.file_validation import validate_excel_upload_with_size_limit

router = APIRouter(
    prefix="/competitors",
    tags=["Competitors"]
)


@router.post("/", response_model=schemas.Competitor, status_code=201)
def create_competitor(
    competitor: schemas.CompetitorCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Create a new competitor for the current user's active brand."""
    return crud.create_competitor(db=db, competitor=competitor, user_id=current_user.id, brand_id=brand_id)


@router.get("/", response_model=List[schemas.Competitor])
def read_competitors(
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Retrieve a list of competitors for the current user's active brand."""
    return crud.get_competitors(db, user_id=current_user.id, brand_id=brand_id, skip=skip, limit=limit)


@router.put("/{competitor_id}", response_model=schemas.Competitor)
def update_competitor(
    competitor_id: int,
    competitor_update: schemas.CompetitorUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a competitor for the current user's active brand."""
    db_competitor = crud.update_competitor(
        db,
        competitor_id=competitor_id,
        competitor_update=competitor_update,
        user_id=current_user.id,
        brand_id=brand_id
    )
    if db_competitor is None:
        raise HTTPException(status_code=404, detail="Competitor not found")
    return db_competitor


@router.delete("/{competitor_id}", response_model=schemas.Competitor)
def delete_competitor(
    competitor_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a competitor for the current user's active brand."""
    deleted_competitor = crud.delete_competitor(db, competitor_id=competitor_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_competitor is None:
        raise HTTPException(status_code=404, detail="Competitor not found")
    return deleted_competitor


@router.post("/upload")
async def upload_competitors(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Upload competitors from an Excel file (admin only). Expected columns: competitor_name, active"""
    if not brand_id:
        raise HTTPException(status_code=400, detail="No active brand selected")

    # Get the brand to determine the owner's user_id
    brand = db.query(models.BrandInfo).filter(models.BrandInfo.id == brand_id).first()
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    # Use the brand owner's user_id, not the admin's user_id
    # This ensures data ownership remains with the brand owner
    owner_user_id = brand.user_id

    try:
        # Validate file type and size using magic byte checking
        contents = await validate_excel_upload_with_size_limit(file)

        # Read the Excel file
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Validate required columns
        required_columns = ['competitor_name']
        for col in required_columns:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        created_count = 0
        updated_count = 0
        errors = []

        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dict from row
                row_data = dict(zip(headers, row))

                if not row_data.get('competitor_name'):
                    continue  # Skip empty rows

                # Check if competitor already exists
                existing_competitor = db.query(models.Competitor).filter(
                    models.Competitor.competitor_name == row_data['competitor_name'],
                    models.Competitor.user_id == owner_user_id,
                    models.Competitor.brand_id == brand_id
                ).first()

                if existing_competitor:
                    # Update existing competitor
                    if 'active' in row_data and row_data['active'] is not None:
                        existing_competitor.active = bool(row_data['active'])
                    updated_count += 1
                else:
                    # Create new competitor
                    new_competitor = models.Competitor(
                        user_id=owner_user_id,  # Use brand owner's ID, not admin's ID
                        brand_id=brand_id,
                        competitor_name=row_data['competitor_name'],
                        active=row_data.get('active', True) if row_data.get('active') is not None else True
                    )
                    db.add(new_competitor)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        db.commit()

        return {
            "message": "Competitors uploaded successfully",
            "created": created_count,
            "updated": updated_count,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
