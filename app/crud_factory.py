"""
Generic CRUD Router Factory
Creates standard CRUD endpoints with minimal configuration, eliminating code duplication.
"""
from typing import Type, Callable, Optional, Dict, Any, List
from dataclasses import dataclass, field
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, status, Path, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel
import io
from openpyxl import load_workbook

from app import models
from app.auth import get_current_user, get_current_admin_user
from app.database import get_db
from app.routers.analytics import get_active_brand_id


@dataclass
class UploadConfig:
    """Configuration for Excel upload endpoint."""
    required_columns: List[str]
    unique_field: str
    field_mapping: Dict[str, str]
    defaults: Dict[str, Any] = field(default_factory=dict)
    admin_only: bool = True


@dataclass
class ListFilter:
    """Configuration for a list endpoint filter parameter."""
    name: str
    type: Type
    default: Any
    crud_function: Optional[Callable] = None  # If provided, uses this instead of default get_list


@dataclass
class CrudConfig:
    """Configuration for CRUD router generation."""

    # Resource identification
    resource_name: str              # "query", "competitor"
    resource_name_plural: str       # "queries", "competitors"
    resource_display_name: str      # "Query", "Competitor"

    # Models and Schemas
    model: Type[Any]                # models.Query
    schema: Type[BaseModel]         # schemas.Query
    schema_create: Type[BaseModel]  # schemas.QueryCreate
    schema_update: Type[BaseModel]  # schemas.QueryUpdate

    # CRUD functions
    crud_create: Callable
    crud_get_list: Callable
    crud_get_by_id: Optional[Callable] = None
    crud_update: Optional[Callable] = None
    crud_delete: Optional[Callable] = None

    # ID configuration
    id_field_name: str = "id"       # "id", "query_id", "descriptor_id"
    id_field_type: Type = int       # int or str

    # Optional list filters
    list_filters: List[ListFilter] = field(default_factory=list)

    # Upload configuration
    upload_config: Optional[UploadConfig] = None

    # Custom validation hooks
    pre_create_hook: Optional[Callable] = None
    post_create_hook: Optional[Callable] = None


def create_crud_router(config: CrudConfig) -> APIRouter:
    """
    Factory function to generate a complete CRUD router with minimal configuration.

    Returns a FastAPI APIRouter with the following endpoints:
    - POST / (create)
    - GET / (list with pagination)
    - GET /{id} (read single) [optional]
    - PUT /{id} (update) [optional]
    - DELETE /{id} (delete) [optional]
    - POST /upload (Excel bulk import) [optional]

    Args:
        config: CrudConfig object with router configuration

    Returns:
        Configured FastAPI APIRouter
    """

    router = APIRouter(
        prefix=f"/{config.resource_name_plural}",
        tags=[config.resource_display_name + "s"]
    )

    # --- CREATE endpoint ---
    @router.post("/", response_model=config.schema, status_code=status.HTTP_201_CREATED)
    async def create_resource(
        resource: config.schema_create,
        current_user: models.User = Depends(get_current_user),
        db: Session = Depends(get_db),
        brand_id: Optional[int] = Depends(get_active_brand_id)
    ):
        f"""Create a new {config.resource_display_name.lower()} for the current user's active brand."""

        # Run pre-create hook if provided (for validation, duplicate checks, etc.)
        if config.pre_create_hook:
            config.pre_create_hook(db, resource, current_user, brand_id)

        # Create the resource
        created = config.crud_create(
            db=db,
            **{config.resource_name: resource},
            user_id=current_user.id,
            brand_id=brand_id
        )

        # Run post-create hook if provided
        if config.post_create_hook:
            config.post_create_hook(db, created, current_user, brand_id)

        return created

    # --- LIST endpoint ---
    # Build list endpoint dynamically based on filters
    if config.list_filters:
        # Create endpoint with filters
        filter_params = {
            f.name: (f.type, Query(f.default)) for f in config.list_filters
        }

        @router.get("/", response_model=List[config.schema])
        async def list_resources(
            skip: int = Query(0),
            limit: int = Query(100),
            current_user: models.User = Depends(get_current_user),
            db: Session = Depends(get_db),
            brand_id: Optional[int] = Depends(get_active_brand_id),
            **filter_params
        ):
            f"""Retrieve a list of {config.resource_name_plural} for the current user's active brand."""

            # Check if any filter has a custom crud function
            for list_filter in config.list_filters:
                if list_filter.name in filter_params:
                    filter_value = filter_params[list_filter.name]
                    if filter_value and list_filter.crud_function:
                        return list_filter.crud_function(
                            db,
                            user_id=current_user.id,
                            brand_id=brand_id,
                            skip=skip,
                            limit=limit
                        )

            # Use default list function
            return config.crud_get_list(
                db,
                user_id=current_user.id,
                brand_id=brand_id,
                skip=skip,
                limit=limit
            )
    else:
        # Simple list endpoint without filters
        @router.get("/", response_model=List[config.schema])
        async def list_resources(
            skip: int = Query(0),
            limit: int = Query(100),
            current_user: models.User = Depends(get_current_user),
            db: Session = Depends(get_db),
            brand_id: Optional[int] = Depends(get_active_brand_id)
        ):
            f"""Retrieve a list of {config.resource_name_plural} for the current user's active brand."""
            return config.crud_get_list(
                db,
                user_id=current_user.id,
                brand_id=brand_id,
                skip=skip,
                limit=limit
            )

    # --- GET BY ID endpoint (optional) ---
    if config.crud_get_by_id:
        # Create endpoint using closure to capture config
        def make_get_endpoint(cfg: CrudConfig):
            async def get_resource_by_id(
                resource_id: cfg.id_field_type = Path(..., alias=cfg.id_field_name),
                current_user: models.User = Depends(get_current_user),
                db: Session = Depends(get_db),
                brand_id: Optional[int] = Depends(get_active_brand_id)
            ):
                result = cfg.crud_get_by_id(
                    db,
                    **{cfg.id_field_name: resource_id},
                    user_id=current_user.id,
                    brand_id=brand_id
                )

                if result is None:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"{cfg.resource_display_name} not found"
                    )

                return result
            return get_resource_by_id

        # Register the endpoint
        router.add_api_route(
            path=f"/{{{config.id_field_name}}}",
            endpoint=make_get_endpoint(config),
            methods=["GET"],
            response_model=config.schema
        )

    # --- UPDATE endpoint (optional) ---
    if config.crud_update:
        update_param_name = f"{config.resource_name}_update"

        # Create endpoint using closure to capture config and param names
        def make_update_endpoint(cfg: CrudConfig, update_key: str):
            async def update_resource(
                resource_id: cfg.id_field_type = Path(..., alias=cfg.id_field_name),
                resource_update: cfg.schema_update = None,
                current_user: models.User = Depends(get_current_user),
                db: Session = Depends(get_db),
                brand_id: Optional[int] = Depends(get_active_brand_id)
            ):
                result = cfg.crud_update(
                    db,
                    **{cfg.id_field_name: resource_id},
                    **{update_key: resource_update},
                    user_id=current_user.id,
                    brand_id=brand_id
                )

                if result is None:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"{cfg.resource_display_name} not found"
                    )

                return result
            return update_resource

        # Register the endpoint
        router.add_api_route(
            path=f"/{{{config.id_field_name}}}",
            endpoint=make_update_endpoint(config, update_param_name),
            methods=["PUT"],
            response_model=config.schema
        )

    # --- DELETE endpoint (optional) ---
    if config.crud_delete:
        # Create endpoint using closure to capture config
        def make_delete_endpoint(cfg: CrudConfig):
            async def delete_resource(
                resource_id: cfg.id_field_type = Path(..., alias=cfg.id_field_name),
                current_user: models.User = Depends(get_current_user),
                db: Session = Depends(get_db),
                brand_id: Optional[int] = Depends(get_active_brand_id)
            ):
                result = cfg.crud_delete(
                    db,
                    **{cfg.id_field_name: resource_id},
                    user_id=current_user.id,
                    brand_id=brand_id
                )

                if result is None:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"{cfg.resource_display_name} not found"
                    )

                return result
            return delete_resource

        # Register the endpoint
        router.add_api_route(
            path=f"/{{{config.id_field_name}}}",
            endpoint=make_delete_endpoint(config),
            methods=["DELETE"],
            response_model=config.schema
        )

    # --- UPLOAD endpoint (optional) ---
    if config.upload_config:
        upload_cfg = config.upload_config

        @router.post("/upload")
        async def upload_resources(
            file: UploadFile = File(...),
            current_user: models.User = Depends(get_current_admin_user) if upload_cfg.admin_only else Depends(get_current_user),
            db: Session = Depends(get_db),
            brand_id: Optional[int] = Depends(get_active_brand_id)
        ):
            f"""Upload {config.resource_name_plural} from an Excel file. Expected columns: {', '.join(upload_cfg.required_columns)}"""

            if not brand_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="No active brand selected"
                )

            if not file.filename.endswith(('.xlsx', '.xls')):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="File must be an Excel file (.xlsx or .xls)"
                )

            try:
                # Read the Excel file
                contents = await file.read()
                wb = load_workbook(io.BytesIO(contents))
                ws = wb.active

                # Get headers from first row
                headers = [cell.value for cell in ws[1]]

                # Validate required columns
                for col in upload_cfg.required_columns:
                    if col not in headers:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"Missing required column: {col}"
                        )

                created_count = 0
                updated_count = 0
                errors = []

                # Process each row (skip header)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Create dict from row
                        row_data = dict(zip(headers, row))

                        # Skip empty rows
                        if not row_data.get(upload_cfg.unique_field):
                            continue

                        # Check if resource already exists
                        existing = db.query(config.model).filter(
                            getattr(config.model, upload_cfg.unique_field) == row_data[upload_cfg.unique_field],
                            config.model.user_id == current_user.id,
                            config.model.brand_id == brand_id
                        ).first()

                        if existing:
                            # Update existing resource
                            for excel_col, model_field in upload_cfg.field_mapping.items():
                                if excel_col in row_data and row_data[excel_col] is not None:
                                    setattr(existing, model_field, row_data[excel_col])
                            updated_count += 1
                        else:
                            # Create new resource
                            new_data = {
                                'user_id': current_user.id,
                                'brand_id': brand_id
                            }

                            # Map Excel columns to model fields
                            for excel_col, model_field in upload_cfg.field_mapping.items():
                                value = row_data.get(excel_col)
                                if value is not None:
                                    new_data[model_field] = value
                                elif model_field in upload_cfg.defaults:
                                    new_data[model_field] = upload_cfg.defaults[model_field]

                            new_resource = config.model(**new_data)
                            db.add(new_resource)
                            created_count += 1

                    except Exception as e:
                        errors.append(f"Row {row_idx}: {str(e)}")

                db.commit()

                return {
                    "message": f"{config.resource_display_name}s uploaded successfully",
                    "created": created_count,
                    "updated": updated_count,
                    "errors": errors if errors else None
                }

            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Failed to process file: {str(e)}"
                )

    return router
